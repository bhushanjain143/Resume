"""Fill Data Engineering Learning Tracker for a 5-year Azure DE interview profile."""
from copy import copy
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

SRC = r"c:\Users\bhushaja\Downloads\Data_Engineering_Learning_Tracker (1).xlsx"
OUT1 = r"c:\Users\bhushaja\Downloads\Data_Engineering_Learning_Tracker_5YOE_Filled.xlsx"
OUT2 = r"C:\Users\bhushaja\Downloads\data engg\Data_Engineering_Learning_Tracker_5YOE_Filled.xlsx"

# (priority, status, started, completed, notes, resource)
# Status: Completed | In Progress | Revise Again
SQL = {
    "SELECT, WHERE, ORDER BY, DISTINCT, LIMIT/TOP": (
        "High", "Completed", "2021-03-01", "2021-06-01",
        "SAY: Filter in WHERE, sort last. DISTINCT is expensive (sort/hash). Prefer TOP/LIMIT after a precise WHERE. Interview trap: WHERE vs HAVING — WHERE before grouping, HAVING after.",
        "https://learn.microsoft.com/sql/t-sql/queries/select-transact-sql",
    ),
    "Filtering: IN, BETWEEN, LIKE, IS NULL": (
        "High", "Completed", "2021-03-01", "2021-06-01",
        "SAY: IS NULL never equals NULL. LIKE '%x' cannot use a normal index. IN vs EXISTS: EXISTS often better for correlated existence checks. BETWEEN is inclusive.",
        "https://learn.microsoft.com/sql/t-sql/queries/predicates",
    ),
    "GROUP BY and HAVING": (
        "High", "Completed", "2021-03-01", "2021-07-01",
        "SAY: SELECT non-aggregates must be in GROUP BY. HAVING filters groups (e.g. HAVING COUNT(*)>1). Interview: find duplicates = GROUP BY key HAVING COUNT(*)>1.",
        "https://learn.microsoft.com/sql/t-sql/queries/select-group-by-transact-sql",
    ),
    "Data types, constraints, basic DDL/DML": (
        "High", "Completed", "2021-03-01", "2021-08-01",
        "SAY: PK/UNIQUE/NOT NULL/CHECK/FK. DDL=structure (CREATE/ALTER), DML=data (INSERT/UPDATE/DELETE/MERGE). Pick VARCHAR vs NVARCHAR, DATETIME2 vs DATETIME.",
        "https://learn.microsoft.com/sql/t-sql/statements/statements",
    ),
    "COUNT, SUM, AVG, MIN, MAX": (
        "High", "Completed", "2021-03-01", "2021-06-01",
        "SAY: COUNT(*) counts rows; COUNT(col) ignores NULLs. AVG ignores NULLs. Wrap with COALESCE if business wants zeros.",
        "https://learn.microsoft.com/sql/t-sql/functions/aggregate-functions-transact-sql",
    ),
    "GROUP BY with multiple columns": (
        "High", "Completed", "2021-04-01", "2021-08-01",
        "SAY: Grain of the result = all GROUP BY columns. Example: sales by region AND product. If you add a column to SELECT, add it to GROUP BY or aggregate it.",
        "https://learn.microsoft.com/sql/t-sql/queries/select-group-by-transact-sql",
    ),
    "ROLLUP, CUBE, GROUPING SETS": (
        "Medium", "In Progress", "2024-01-15", None,
        "SAY: ROLLUP = hierarchy totals (region→country→grand). CUBE = all combinations. GROUPING SETS = exact totals you want (cheaper than CUBE). GROUPING() flags super-aggregate NULL vs real NULL.",
        "https://learn.microsoft.com/sql/t-sql/queries/select-group-by-transact-sql",
    ),
    "INNER JOIN": (
        "High", "Completed", "2021-03-01", "2021-06-01",
        "SAY: Only matching keys. Use when both sides must exist (order + paying customer). Watch fan-out: join to a 1:M dim without grain control duplicates facts.",
        "https://learn.microsoft.com/sql/t-sql/queries/from-transact-sql",
    ),
    "LEFT / RIGHT JOIN": (
        "High", "Completed", "2021-03-01", "2021-07-01",
        "SAY: LEFT keeps all left rows; unmatched right cols NULL. Filter on right table in WHERE turns it into INNER — put those predicates in ON. RIGHT is just flipped LEFT; I almost always write LEFT.",
        "https://learn.microsoft.com/sql/t-sql/queries/from-transact-sql",
    ),
    "FULL OUTER JOIN": (
        "Medium", "Completed", "2021-08-01", "2022-03-01",
        "SAY: Keep unmatched from both sides. Use for reconciliation (source vs target row counts/keys). Rare in star-schema queries.",
        "https://learn.microsoft.com/sql/t-sql/queries/from-transact-sql",
    ),
    "SELF JOIN": (
        "High", "Completed", "2021-08-01", "2022-03-01",
        "SAY: Table joined to itself with aliases. Classic: employee→manager (e.manager_id = m.emp_id), consecutive days, compare row to previous version. Need aliases or the query is invalid.",
        "https://learn.microsoft.com/sql/t-sql/queries/from-transact-sql",
    ),
    "CROSS JOIN": (
        "Medium", "Completed", "2021-08-01", "2022-03-01",
        "SAY: Cartesian product. Use for generating calendars, all product×region combinations, or numbers tables. Accidental CROSS JOIN is a common explosion bug.",
        "https://learn.microsoft.com/sql/t-sql/queries/from-transact-sql",
    ),
    "Anti-join / Semi-join patterns (NOT EXISTS, NOT IN)": (
        "High", "Completed", "2022-01-01", "2023-06-01",
        "SAY: Semi-join = EXISTS (rows that have a match). Anti-join = NOT EXISTS (no match). Prefer NOT EXISTS over NOT IN because NOT IN fails if the subquery has NULL. Interview favorite.",
        "https://learn.microsoft.com/sql/t-sql/language-elements/exists-transact-sql",
    ),
    "ROW_NUMBER, RANK, DENSE_RANK": (
        "High", "Completed", "2022-01-01", "2023-01-01",
        "SAY: All need OVER(PARTITION BY … ORDER BY …). ROW_NUMBER unique 1,2,3… RANK skips after ties (1,1,3). DENSE_RANK no gaps (1,1,2). Dedup: ROW_NUMBER()=1. Top-N per group: ROW_NUMBER<=N.",
        "https://learn.microsoft.com/sql/t-sql/functions/ranking-functions-transact-sql",
    ),
    "LEAD, LAG": (
        "High", "Completed", "2022-06-01", "2023-06-01",
        "SAY: LAG = previous row, LEAD = next, same partition/order. Use for YoY, sessionization, SCD change detection (current vs previous status). Default offset 1; can supply default for first/last row.",
        "https://learn.microsoft.com/sql/t-sql/functions/analytic-functions-transact-sql",
    ),
    "SUM/AVG OVER (PARTITION BY ... ORDER BY ...)": (
        "High", "Completed", "2022-06-01", "2023-06-01",
        "SAY: Window aggregates without collapsing rows. Running total: SUM(amt) OVER (PARTITION BY cust ORDER BY dt ROWS UNBOUNDED PRECEDING). Default RANGE vs ROWS can surprise with ties — I specify ROWS.",
        "https://learn.microsoft.com/sql/t-sql/queries/select-over-clause-transact-sql",
    ),
    "Running totals & moving averages": (
        "High", "Completed", "2022-06-01", "2024-01-01",
        "SAY: Moving avg 7 days: AVG(x) OVER (ORDER BY d ROWS BETWEEN 6 PRECEDING AND CURRENT ROW). In Spark same idea with window spec. Call out frame clause in interviews.",
        "https://learn.microsoft.com/sql/t-sql/queries/select-over-clause-transact-sql",
    ),
    "NTILE, FIRST_VALUE, LAST_VALUE": (
        "Medium", "Completed", "2023-01-01", "2024-06-01",
        "SAY: NTILE(4) quartiles. FIRST_VALUE/LAST_VALUE need a frame; LAST_VALUE often needs ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING or it returns current row.",
        "https://learn.microsoft.com/sql/t-sql/functions/analytic-functions-transact-sql",
    ),
    "Reading execution/query plans": (
        "High", "Completed", "2022-01-01", "2024-12-01",
        "SAY: Look for Clustered Index Scan vs Seek, Key Lookup, Hash vs Nested Loops, Sort, high estimated vs actual rows. SSMS: Actual plan. Spark: EXPLAIN / Spark UI DAG. Cardinality mistakes cause bad joins.",
        "https://learn.microsoft.com/sql/relational-databases/performance/execution-plans",
    ),
    "Index usage basics": (
        "High", "Completed", "2021-08-01", "2023-06-01",
        "SAY: B-tree index = sorted pointer. Helps WHERE/JOIN/ORDER BY on leading columns. Hurts heavy writes. Covering index includes selected columns to avoid lookups.",
        "https://learn.microsoft.com/sql/relational-databases/indexes/clustered-and-nonclustered-indexes-described",
    ),
    "Avoiding SELECT *, sargable predicates": (
        "High", "Completed", "2022-01-01", "2023-06-01",
        "SAY: Sargable = predicate can use index. Bad: WHERE YEAR(dt)=2024, WHERE col LIKE '%x', WHERE CAST(col AS date)=… Good: WHERE dt>='2024-01-01' AND dt<'2025-01-01'. SELECT * blocks covering indexes and wastes I/O.",
        "https://learn.microsoft.com/sql/relational-databases/performance/query-hints-transact-sql",
    ),
    "CTE vs subquery vs temp table performance": (
        "High", "Completed", "2022-06-01", "2024-06-01",
        "SAY: CTE is named query; may be inlined (not always materialized). Subquery similar. Temp table (#t) materializes, can index, good for reuse/large intermediates. Table variable weak stats. I use CTE for readability; temp table when the optimizer misestimates.",
        "https://learn.microsoft.com/sql/t-sql/queries/with-common-table-expression-transact-sql",
    ),
    "SCD Type 0, 1, 2, 3 concepts": (
        "High", "Completed", "2022-01-01", "2023-06-01",
        "SAY: 0=retain original (audit). 1=overwrite (no history). 2=new row + effective dates/is_current (full history) — most used in DWH. 3=previous-value column (limited history). Type 6 = 1+2+3 hybrid.",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/type-2/",
    ),
    "Implementing SCD Type 2 with SQL": (
        "High", "Completed", "2022-06-01", "2024-01-01",
        "SAY: MERGE source to target: 1) expire current row (is_current=0, end_date=today) when attributes change 2) insert new current row with new surrogate key. Hash of business attrs to detect change. Always mention surrogate key.",
        "https://learn.microsoft.com/sql/t-sql/statements/merge-transact-sql",
    ),
    "Change Data Capture concepts": (
        "High", "Completed", "2022-06-01", "2024-06-01",
        "SAY: Capture inserts/updates/deletes from source instead of full dump. Enables incremental + deletes. Sources: SQL CDC, Oracle GoldenGate, Debezium, ADF watermark, Databricks Auto Loader. Need a reliable change identifier.",
        "https://learn.microsoft.com/sql/relational-databases/track-changes/about-change-data-capture-sql-server",
    ),
    "CDC implementation patterns (timestamp/log-based)": (
        "High", "Completed", "2023-01-01", "2025-06-01",
        "SAY: Timestamp/watermark: WHERE updated_at > last_success (misses deletes unless soft-delete). Log-based CDC: read transaction log (true deletes, lower source load). I stored last watermark in a control table and advanced it only after successful load.",
        "https://learn.microsoft.com/azure/data-factory/tutorial-incremental-copy-overview",
    ),
    "EXISTS vs IN vs JOIN performance": (
        "High", "Completed", "2022-01-01", "2024-01-01",
        "SAY: EXISTS short-circuits on first match (semi-join). IN builds a set; NULL in list breaks NOT IN. JOIN can duplicate if 1:M. For 'does a child exist?' I use EXISTS. Confirm with actual plan.",
        "https://learn.microsoft.com/sql/t-sql/language-elements/exists-transact-sql",
    ),
    "MERGE statement (upsert logic)": (
        "High", "Completed", "2022-06-01", "2024-01-01",
        "SAY: WHEN MATCHED THEN UPDATE / WHEN NOT MATCHED THEN INSERT / WHEN NOT MATCHED BY SOURCE THEN expire. Used for SCD1/2 and Delta MERGE. Caution: duplicate source keys fail MERGE; holdlock/isolation; Spark MERGE is the lake equivalent.",
        "https://learn.microsoft.com/sql/t-sql/statements/merge-transact-sql",
    ),
    "Correlated vs non-correlated subqueries": (
        "High", "Completed", "2022-01-01", "2023-06-01",
        "SAY: Non-correlated runs once. Correlated runs per outer row (like a loop) — rewrite as JOIN/window if slow. Example correlated: WHERE salary > (SELECT AVG(salary) FROM emp e2 WHERE e2.dept=e1.dept) — better as window AVG.",
        "https://learn.microsoft.com/sql/t-sql/queries/subqueries",
    ),
    "Common Table Expressions (CTE)": (
        "High", "Completed", "2021-08-01", "2022-12-01",
        "SAY: WITH cte AS (SELECT…) SELECT… Improves readability, can chain. Not a stored object. INSERT/UPDATE/DELETE can target a CTE if it maps to one base table. Recursion is a special CTE.",
        "https://learn.microsoft.com/sql/t-sql/queries/with-common-table-expression-transact-sql",
    ),
    "Recursive CTE": (
        "Medium", "In Progress", "2023-06-01", None,
        "SAY: Anchor + UNION ALL recursive member referencing the CTE. Use for org charts, bill-of-materials, graph walks. Always a terminator (level < N) to avoid infinite loops. OPTION (MAXRECURSION n).",
        "https://learn.microsoft.com/sql/t-sql/queries/with-common-table-expression-transact-sql",
    ),
    "Clustered vs Non-clustered index": (
        "High", "Completed", "2021-08-01", "2023-01-01",
        "SAY: Clustered = table sorted by key (one per table, often PK). Nonclustered = extra B-tree pointing to clustered key/RID. Lookups happen if NC index is not covering. Heap = no clustered index.",
        "https://learn.microsoft.com/sql/relational-databases/indexes/clustered-and-nonclustered-indexes-described",
    ),
    "When indexes help vs hurt performance": (
        "High", "Completed", "2022-01-01", "2024-06-01",
        "SAY: Help: selective filters, joins, sorts. Hurt: high-write tables, low-selectivity columns (gender), too many indexes slowing INSERT. ETL: sometimes drop/disable NC indexes, load, rebuild. Over-indexing is a real cost.",
        "https://learn.microsoft.com/sql/relational-databases/indexes/indexes",
    ),
    "Normalization (1NF, 2NF, 3NF)": (
        "High", "Completed", "2021-03-01", "2022-06-01",
        "SAY: 1NF atomic values. 2NF no partial key dependency. 3NF no transitive dependency. OLTP loves 3NF. DWH Gold often denormalizes dims for query speed.",
        "https://learn.microsoft.com/office/troubleshoot/access/database-normalization-description",
    ),
    "Denormalization trade-offs": (
        "High", "Completed", "2022-01-01", "2023-06-01",
        "SAY: Fewer joins, faster BI, but redundancy and extra ETL to keep in sync. Star schema is controlled denormalization of dimensions.",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/",
    ),
    "ACID properties & isolation levels": (
        "High", "Completed", "2022-01-01", "2024-01-01",
        "SAY: Atomic, Consistent, Isolated, Durable. Levels: Read Uncommitted < Read Committed (default SQL Server) < Repeatable Read < Serializable; Snapshot uses row versions. Delta Lake gives ACID on the lake via _delta_log.",
        "https://learn.microsoft.com/sql/t-sql/statements/set-transaction-isolation-level-transact-sql",
    ),
    "Locking & deadlocks basics": (
        "Medium", "Completed", "2023-01-01", "2025-01-01",
        "SAY: Deadlock = two sessions wait on each other; engine kills one (victim). Fix: consistent lock order, shorter tx, proper indexes, READ COMMITTED SNAPSHOT. In Spark, concurrent writes need Delta isolation, not DB locks.",
        "https://learn.microsoft.com/sql/relational-databases/sql-server-transaction-locking-and-row-versioning-guide",
    ),
    "Views, Stored Procedures, Triggers (basics)": (
        "High", "Completed", "2021-06-01", "2023-01-01",
        "SAY: View=saved SELECT (security/abstraction). SP=procedural logic, parameters, good for ETL in SQL. Trigger=fires on DML — I avoid heavy triggers in warehouses (hidden, hard to debug). Indexed/materialized views when needed.",
        "https://learn.microsoft.com/sql/relational-databases/stored-procedures/stored-procedures-database-engine",
    ),
    "PIVOT / UNPIVOT": (
        "Medium", "In Progress", "2023-06-01", None,
        "SAY: PIVOT turns rows to columns (months as columns). UNPIVOT opposite. In Spark: pivot() on DataFrame or stack/expr. For many dynamic values I often use CASE WHEN instead of PIVOT.",
        "https://learn.microsoft.com/sql/t-sql/queries/from-using-pivot-and-unpivot",
    ),
    "Common string & date functions": (
        "High", "Completed", "2021-03-01", "2022-01-01",
        "SAY: CONCAT, SUBSTRING, REPLACE, LTRIM; CAST/CONVERT, DATEADD, DATEDIFF, DATETRUNC, FORMAT (slow). Always store UTC and convert in Gold/BI. Spark: to_date, date_trunc, concat_ws.",
        "https://learn.microsoft.com/sql/t-sql/functions/date-and-time-data-types-and-functions-transact-sql",
    ),
    "Solve 5-10 SQL queries daily (track in Daily SQL Log tab)": (
        "High", "In Progress", "2026-08-24", None,
        "DO: 5-10 queries/day on LeetCode SQL / HackerRank / StrataScratch. Rotate: joins, windows, gaps-and-islands, SCD, optimization. Log in Daily SQL Practice Log. At 5 YOE they expect speed + correctness.",
        "https://leetcode.com/problemset/database/",
    ),
}

DWH = {
    "Fact table concepts (grain, measures)": (
        "High", "Completed", "2021-09-01", "2022-06-01",
        "SAY: Grain first — one row = one business event (e.g. one order line). Measures: additive (amount), semi-additive (balance), non-additive (ratio). Never mix grains in one fact.",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/fact-table/",
    ),
    "Dimension table concepts (attributes, hierarchies)": (
        "High", "Completed", "2021-09-01", "2022-06-01",
        "SAY: Dims describe the who/what/where/when. Hierarchies: Product→Category→Dept, Date→Month→Year. Wide, denormalized in star. Surrogate integer keys join to facts.",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/dimension-table-structure/",
    ),
    "Types of fact tables (transactional, snapshot, accumulating)": (
        "High", "Completed", "2022-01-01", "2023-06-01",
        "SAY: Transaction = one row per event. Periodic snapshot = inventory/balance each day. Accumulating snapshot = process pipeline (order→ship→deliver) with multiple date keys and milestones.",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/transaction-fact-table/",
    ),
    "Star Schema": (
        "High", "Completed", "2021-09-01", "2022-06-01",
        "SAY: Central fact + denormalized dims. Simple joins, BI-friendly, my default for Gold. Draw: FactSales in center, DimDate/Customer/Product/Store around it.",
        "https://learn.microsoft.com/power-bi/guidance/star-schema",
    ),
    "Snowflake Schema": (
        "Medium", "Completed", "2022-01-01", "2023-01-01",
        "SAY: Dims normalized into sub-dims (Product→Category table). Saves space, more joins. I snowflake only when a dim is huge or shared in a strict 3NF warehouse.",
        "https://learn.microsoft.com/power-bi/guidance/star-schema",
    ),
    "Star vs Snowflake trade-offs": (
        "High", "Completed", "2022-01-01", "2023-06-01",
        "SAY: Star = faster queries, simpler Power BI. Snowflake = less duplication, more joins, harder for BI. Interview: Power BI prefers star.",
        "https://learn.microsoft.com/power-bi/guidance/star-schema",
    ),
    "Primary Key": (
        "High", "Completed", "2021-03-01", "2021-08-01",
        "SAY: Uniquely identifies a row; implies NOT NULL + UNIQUE; in SQL Server usually clustered. One PK per table. In lakehouse, PK may be declared in Unity Catalog but not always physically enforced like RDBMS.",
        "https://learn.microsoft.com/sql/relational-databases/tables/primary-and-foreign-key-constraints",
    ),
    "Foreign Key": (
        "High", "Completed", "2021-03-01", "2021-08-01",
        "SAY: Child column references parent PK/UNIQUE. Enforces referential integrity. In warehouse ETL we often enforce in pipeline (orphan checks) rather than heavy FKs on large facts.",
        "https://learn.microsoft.com/sql/relational-databases/tables/primary-and-foreign-key-constraints",
    ),
    "Surrogate Key vs Natural Key": (
        "High", "Completed", "2021-09-01", "2022-12-01",
        "SAY: Natural = business key (email, product SKU) — can change, composite, wide. Surrogate = warehouse-generated integer/hash — stable for SCD2 (new version = new SK). Facts store surrogate keys.",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/surrogate-keys/",
    ),
    "Composite Key": (
        "Medium", "Completed", "2021-06-01", "2022-06-01",
        "SAY: PK of 2+ columns (order_id+line_no). Fine in OLTP. In DWH I still add a surrogate for facts/dims to keep joins skinny.",
        "https://learn.microsoft.com/sql/relational-databases/tables/primary-and-foreign-key-constraints",
    ),
    "OLTP characteristics": (
        "High", "Completed", "2021-03-01", "2022-01-01",
        "SAY: Many small read/write tx, normalized, current state, app-facing (orders, CRM). Latency in ms. Azure SQL / on-prem SQL Server.",
        "https://learn.microsoft.com/azure/architecture/data-guide/relational-data/online-transaction-processing",
    ),
    "OLAP characteristics": (
        "High", "Completed", "2021-09-01", "2022-06-01",
        "SAY: Analytical, large scans, historical, denormalized/star, aggregations. Synapse, Fabric Warehouse, Databricks SQL, Power BI. Latency seconds to minutes.",
        "https://learn.microsoft.com/azure/architecture/data-guide/relational-data/online-analytical-processing",
    ),
    "Use case differences": (
        "High", "Completed", "2021-09-01", "2022-06-01",
        "SAY: Checkout system = OLTP. Monthly sales dashboard / what-if = OLAP. Don't run heavy reports on OLTP; replicate/CDC to the warehouse.",
        "https://learn.microsoft.com/azure/architecture/data-guide/relational-data/online-analytical-processing",
    ),
    "ETL vs ELT": (
        "High", "Completed", "2021-09-01", "2023-01-01",
        "SAY: ETL transform in engine then load (SSIS/ADF mapping). ELT land raw then transform in lake/warehouse (ADF copy → Databricks/Synapse). Cloud default is ELT + medallion.",
        "https://learn.microsoft.com/azure/architecture/data-guide/relational-data/etl",
    ),
    "Full Load": (
        "High", "Completed", "2021-06-01", "2022-06-01",
        "SAY: Truncate+reload or overwrite folder. Simple, handles deletes, expensive. Use for small dims or first load. Staging then swap for atomicity.",
        "https://learn.microsoft.com/azure/data-factory/tutorial-incremental-copy-overview",
    ),
    "Incremental Load": (
        "High", "Completed", "2022-01-01", "2023-06-01",
        "SAY: Load only new/changed rows via watermark, CDC, or hash compare. Faster, cheaper. Must define delete strategy (soft-delete flag or CDC).",
        "https://learn.microsoft.com/azure/data-factory/tutorial-incremental-copy-overview",
    ),
    "Batch Processing mode": (
        "High", "Completed", "2021-06-01", "2023-01-01",
        "SAY: Bounded data on a schedule (hourly/daily ADF trigger + Databricks job). High throughput, simpler exactly-once. SLA = batch window.",
        "https://learn.microsoft.com/azure/architecture/data-guide/big-data/batch-processing",
    ),
    "Streaming mode": (
        "High", "Completed", "2023-01-01", "2025-06-01",
        "SAY: Unbounded events (Event Hub/Kafka → Structured Streaming / Auto Loader). Watermarks for late data. Micro-batch is the usual Spark mode. I used it for near-real-time Bronze.",
        "https://learn.microsoft.com/azure/architecture/data-guide/big-data/real-time-processing",
    ),
    "Kimball (bottom-up) vs Inmon (top-down)": (
        "Medium", "Completed", "2022-06-01", "2024-01-01",
        "SAY: Kimball: dimensional marts first, conformed dims, faster delivery — what I use. Inmon: normalized enterprise EDW first, then marts — heavier governance. Hybrid is common.",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/",
    ),
    "Conceptual, Logical, Physical data models": (
        "Medium", "Completed", "2022-01-01", "2023-06-01",
        "SAY: Conceptual=business entities. Logical=keys/attributes/relationships, platform-agnostic. Physical=tables, partitions, file format, indexes. Interviews: I start logical grain then physical Delta partition strategy.",
        "https://learn.microsoft.com/power-bi/guidance/star-schema",
    ),
    "Data Mart vs Data Warehouse": (
        "High", "Completed", "2021-09-01", "2022-12-01",
        "SAY: Warehouse = enterprise, many subjects. Mart = department slice (Finance/Sales) often star. Gold layer marts feed Power BI.",
        "https://learn.microsoft.com/azure/architecture/data-guide/relational-data/data-warehousing",
    ),
    "Conformed Dimension": (
        "High", "Completed", "2022-06-01", "2024-01-01",
        "SAY: Same dim reused across facts/marts (DimDate, DimCustomer) so reports can drill across processes. Mastered once in Silver/Gold.",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/conformed-dimension/",
    ),
    "Junk Dimension": (
        "Medium", "Completed", "2023-01-01", "2024-06-01",
        "SAY: Pack low-cardinality flags/codes (yes/no, status) into one dim to avoid dozens of tiny dims. Degenerate is different (see next).",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/junk-dimension/",
    ),
    "Degenerate Dimension": (
        "Medium", "Completed", "2023-01-01", "2024-06-01",
        "SAY: A dimension attribute stored on the fact because it has no other attributes — e.g. invoice_number, ticket_id. No separate dim table.",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/degenerate-dimension/",
    ),
    "Slowly Changing Dimension Type 3, 4, 6 (hybrid)": (
        "Medium", "In Progress", "2024-01-01", None,
        "SAY: T3=current + previous column. T4=history mini-table (mini-dimension). T6=1+2+3 (current overwrite + history rows + previous). I implement T2 in production; mention T6 if they ask hybrids.",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/type-6/",
    ),
    "Factless Fact Table": (
        "Medium", "Completed", "2023-01-01", "2024-06-01",
        "SAY: No numeric measure — just events or coverage (student attended class, promo was in effect). COUNT(*) is the measure. Used for many-to-many coverage.",
        "https://www.kimballgroup.com/data-warehouse-business-intelligence-resources/kimball-techniques/dimensional-modeling-techniques/factless-fact-table/",
    ),
}

PYSPARK = {
    "Driver, Executor, Cluster Manager overview": (
        "High", "Completed", "2022-01-01", "2023-06-01",
        "SAY: Driver builds DAG and schedules. Executors run tasks and cache data. Cluster manager (YARN/K8s/Databricks) grants CPUs/RAM. If driver OOMs, you collected too much to driver.",
        "https://spark.apache.org/docs/latest/cluster-overview.html",
    ),
    "How an application runs in the backend (submission to execution)": (
        "High", "Completed", "2022-06-01", "2024-01-01",
        "SAY: Submit job → cluster starts → SparkSession → lazy transformations build DAG → action triggers jobs. Databricks job cluster: VM provision → Spark start → notebook → teardown.",
        "https://spark.apache.org/docs/latest/job-scheduling.html",
    ),
    "Job -> Stages -> Tasks breakdown": (
        "High", "Completed", "2022-06-01", "2024-01-01",
        "SAY: Action = Job. Shuffle boundary = new Stage. Task = one partition on one core. 200 shuffle partitions default — tune spark.sql.shuffle.partitions / AQE.",
        "https://spark.apache.org/docs/latest/cluster-overview.html",
    ),
    "Lazy Evaluation & DAG": (
        "High", "Completed", "2022-01-01", "2023-01-01",
        "SAY: Transformations don't run until an action (write, count, show). DAG is lineage so Spark can recompute lost partitions. Explain() shows the plan before you pay for a job.",
        "https://spark.apache.org/docs/latest/rdd-programming-guide.html#rdd-operations",
    ),
    "CSV": (
        "High", "Completed", "2021-06-01", "2022-01-01",
        "SAY: Row text, no types, no pushdown. Landing only. InferSchema scans twice. I set header, delimiter, quote, mode=PERMISSIVE/FAILFAST.",
        "https://spark.apache.org/docs/latest/sql-data-sources-csv.html",
    ),
    "Parquet": (
        "High", "Completed", "2022-01-01", "2023-01-01",
        "SAY: Columnar, compressed, predicate/column pushdown. Default analytics format. Delta = Parquet + _delta_log. Partition on low-cardinality columns used in filters.",
        "https://spark.apache.org/docs/latest/sql-data-sources-parquet.html",
    ),
    "AVRO": (
        "Medium", "Completed", "2023-01-01", "2024-06-01",
        "SAY: Row-based, strong schema, good for Kafka CDC. Heavier for analytics than Parquet. I convert Avro/JSON Bronze → Parquet/Delta Silver.",
        "https://spark.apache.org/docs/latest/sql-data-sources-avro.html",
    ),
    "Creating & transforming DataFrames": (
        "High", "Completed", "2022-01-01", "2023-01-01",
        "SAY: spark.read / createDataFrame / spark.sql. Immutable: each transform returns a new DF. Prefer built-in functions over UDFs.",
        "https://spark.apache.org/docs/latest/sql-getting-started.html",
    ),
    "Common transformations (select, filter, withColumn, groupBy)": (
        "High", "Completed", "2022-01-01", "2023-01-01",
        "SAY: Narrow: select/filter/withColumn. Wide: groupBy/join/distinct. Chain transforms; one action at the end. withColumn in a loop is an anti-pattern — use select with expressions.",
        "https://spark.apache.org/docs/latest/api/python/reference/pyspark.sql/dataframe.html",
    ),
    "Actions vs Transformations": (
        "High", "Completed", "2022-01-01", "2023-01-01",
        "SAY: Transform=lazy (map/filter/select). Action=eager (count, collect, write, show). Multiple actions recompute unless cache. collect() on big data kills the driver.",
        "https://spark.apache.org/docs/latest/rdd-programming-guide.html#rdd-operations",
    ),
    "Broadcast Hash Join": (
        "High", "Completed", "2023-01-01", "2024-06-01",
        "SAY: Small side (dim) copied to every executor; no shuffle of the big fact. Hint: broadcast(df). Threshold spark.sql.autoBroadcastJoinThreshold (~10MB default, raise carefully). Best 5-YOE join answer.",
        "https://spark.apache.org/docs/latest/sql-performance-tuning.html#broadcast-hash-join",
    ),
    "Sort Merge Join": (
        "High", "Completed", "2023-01-01", "2024-06-01",
        "SAY: Default for two large tables: shuffle + sort both by key, merge. Needs even keys. Skew on one key blows one task. AQE can split skewed partitions.",
        "https://spark.apache.org/docs/latest/sql-performance-tuning.html",
    ),
    "Shuffle Hash Join": (
        "Medium", "In Progress", "2024-06-01", None,
        "SAY: Shuffle both sides, hash-join per partition (no full sort). Used when tables aren't huge enough for SMJ but too big to broadcast. Less common to tune manually; know it exists vs BHJ/SMJ.",
        "https://www.databricks.com/blog/2020/12/05/faster-sql-adaptive-query-execution-in-databricks.html",
    ),
    "Partitioning": (
        "High", "Completed", "2022-06-01", "2024-01-01",
        "SAY: Spark partitions = parallelism. Table partitions = directory layout (date=). Too many small partitions = small-file problem. Partition by ingestion_date or high-filter columns with sane cardinality.",
        "https://spark.apache.org/docs/latest/sql-data-sources-load-save-functions.html#partition-discovery",
    ),
    "Repartition vs Coalesce": (
        "High", "Completed", "2023-01-01", "2024-06-01",
        "SAY: repartition(n) full shuffle, can increase or decrease, good before writes. coalesce(n) only shrinks, no full shuffle, may leave skew. Before write I often repartition to target file count.",
        "https://spark.apache.org/docs/latest/api/python/reference/pyspark.sql/api/pyspark.sql.DataFrame.repartition.html",
    ),
    "Shuffle operations": (
        "High", "Completed", "2022-06-01", "2024-01-01",
        "SAY: Network exchange across executors (join/groupBy/repartition). Dominant cost. Reduce shuffles, increase parallelism, fix skew, enable AQE. Spark UI: Shuffle Read/Write bytes.",
        "https://spark.apache.org/docs/latest/rdd-programming-guide.html#shuffle-operations",
    ),
    "Cardinality": (
        "High", "Completed", "2023-01-01", "2024-06-01",
        "SAY: Distinct count of a key. High cardinality partition columns = millions of folders (bad). Low cardinality join key can skew. Estimate with approx_count_distinct.",
        "https://spark.apache.org/docs/latest/sql-ref-functions-builtin.html",
    ),
    "Data Skew & handling techniques (salting)": (
        "High", "Completed", "2023-06-01", "2025-06-01",
        "SAY: One key (NULL, 'Unknown', popular customer) owns most data → one slow task. Fixes: AQE skew join, filter nulls, salt key (key+random) then aggregate, two-phase aggregate, isolate hot keys.",
        "https://docs.databricks.com/en/optimizations/aqe.html",
    ),
    "Broadcast variables/joins": (
        "High", "Completed", "2023-01-01", "2024-06-01",
        "SAY: Broadcast join = DF hint. Broadcast variable = read-only lookup (dict) shipped once to executors for UDFs/maps. Don't broadcast a huge DF (driver/executor OOM).",
        "https://spark.apache.org/docs/latest/rdd-programming-guide.html#broadcast-variables",
    ),
    "Predicate Pushdown": (
        "High", "Completed", "2023-01-01", "2024-06-01",
        "SAY: Filter applied at Parquet/Delta scan so we read fewer row groups. Needs filter on physical columns, not wrapped in UDF. Partition pruning is the folder-level version.",
        "https://spark.apache.org/docs/latest/sql-data-sources-parquet.html",
    ),
    "Caching vs Persistence (storage levels)": (
        "Medium", "Completed", "2023-06-01", "2025-01-01",
        "SAY: cache() = MEMORY_AND_DISK. persist(MEMORY_ONLY) faster but can OOM. Use only if reused across actions. unpersist() after. Delta tables usually cheaper to re-read than cache huge DFs.",
        "https://spark.apache.org/docs/latest/rdd-programming-guide.html#rdd-persistence",
    ),
    "Serialization (Kryo vs Java)": (
        "Low", "In Progress", "2024-06-01", None,
        "SAY: Kryo is faster/smaller than Java serializer for RDDs. DataFrames already use Tungsten binary. Mention Kryo if they ask RDD tuning; not the first lever for DataFrame jobs.",
        "https://spark.apache.org/docs/latest/tuning.html#data-serialization",
    ),
    "Adaptive Query Execution (AQE)": (
        "High", "Completed", "2024-01-01", "2025-12-01",
        "SAY: Spark 3+ re-plans at runtime: coalesces shuffle partitions, switches to broadcast if size small, handles skew. spark.sql.adaptive.enabled=true (default on Databricks). Check Spark UI for AQE events.",
        "https://spark.apache.org/docs/latest/sql-performance-tuning.html#adaptive-query-execution",
    ),
    "Small file problem & solutions": (
        "High", "Completed", "2023-01-01", "2025-06-01",
        "SAY: Thousands of tiny Parquet files kill listing and jobs. Causes: over-partitioning, many parallel writers. Fixes: coalesce/repartition before write, Auto Loader + OPTIMIZE, target file size 128–1024MB, liquid clustering.",
        "https://docs.databricks.com/en/delta/best-practices.html",
    ),
    "RDD vs DataFrame vs Dataset": (
        "High", "Completed", "2022-01-01", "2023-06-01",
        "SAY: RDD=untyped partitions, no optimizer. DataFrame=Dataset[Row], Catalyst+Tungsten — what I use in PySpark. Dataset=typed JVM API, not in Python. Prefer DataFrame functions.",
        "https://spark.apache.org/docs/latest/sql-programming-guide.html",
    ),
    "Spark SQL basics": (
        "High", "Completed", "2022-01-01", "2023-06-01",
        "SAY: df.createOrReplaceTempView; spark.sql. Same engine as DataFrame API. Good for analysts; I mix both. Unity Catalog tables: spark.table('cat.schema.t').",
        "https://spark.apache.org/docs/latest/sql-programming-guide.html",
    ),
    "UDFs & Pandas UDFs (and their performance cost)": (
        "High", "Completed", "2023-06-01", "2025-06-01",
        "SAY: Python UDF row-at-a-time, kills Catalyst, serialization cost. Pandas UDF (vectorized Arrow) better. Best: native Spark functions. I only UDF when no built-in exists.",
        "https://spark.apache.org/docs/latest/api/python/user_guide/sql/arrow_pandas.html",
    ),
    "Cluster deployment modes (client vs cluster)": (
        "Medium", "Completed", "2023-06-01", "2025-01-01",
        "SAY: Client: driver on submit machine (notebooks). Cluster: driver on a worker (production jobs). Databricks abstracts this; job cluster vs all-purpose is the practical distinction.",
        "https://spark.apache.org/docs/latest/cluster-overview.html",
    ),
    "Memory management (execution vs storage memory, spill)": (
        "Medium", "Completed", "2024-01-01", "2025-12-01",
        "SAY: Unified memory: execution (shuffles/joins) vs storage (cache). Spill to disk if RAM short — slow but not fail. OOM = raise worker memory, reduce partition size, avoid cache+shuffle together.",
        "https://spark.apache.org/docs/latest/tuning.html#memory-management-overview",
    ),
    "Checkpointing": (
        "Medium", "In Progress", "2024-06-01", None,
        "SAY: Truncate RDD lineage to reliable storage. Streaming: checkpoint location for offsets/state. Delta streaming needs a checkpoint path per stream. Don't share checkpoints across pipelines.",
        "https://spark.apache.org/docs/latest/structured-streaming-programming-guide.html#recovering-from-failures-with-checkpointing",
    ),
    "Reading the Spark UI (DAG, stages, tasks, shuffle)": (
        "High", "Completed", "2023-01-01", "2025-06-01",
        "SAY: Jobs→Stages: look for stragglers, GC, shuffle spill, input vs output. SQL tab for plans. One slow task = skew. Many tiny tasks = too many partitions.",
        "https://spark.apache.org/docs/latest/web-ui.html",
    ),
    "Structured Streaming fundamentals (micro-batch vs continuous)": (
        "High", "Completed", "2023-06-01", "2025-06-01",
        "SAY: Default micro-batch: trigger every N seconds, exactly-once with checkpoint+Delta. Continuous processing is low-latency experimental. Auto Loader is streaming file ingest.",
        "https://spark.apache.org/docs/latest/structured-streaming-programming-guide.html",
    ),
    "Watermarking & windowed aggregations": (
        "High", "In Progress", "2024-06-01", None,
        "SAY: Watermark = max event-time lateness allowed before a window is closed. withWatermark('ts','1 hour') + groupBy(window(ts,'5 minutes')). Without watermark, state grows forever.",
        "https://spark.apache.org/docs/latest/structured-streaming-programming-guide.html#window-operations-on-event-time",
    ),
}

DBX = {
    "Delta Lake fundamentals (transaction log, ACID)": (
        "High", "Completed", "2023-01-01", "2024-06-01",
        "SAY: Open table format: Parquet data + _delta_log JSON/checkpoint. ACID, MERGE, time travel. Concurrent writers coordinated by log. This is the core of a Databricks lakehouse.",
        "https://docs.databricks.com/en/delta/index.html",
    ),
    "Time Travel & versioning": (
        "High", "Completed", "2023-06-01", "2024-12-01",
        "SAY: VERSION AS OF n or TIMESTAMP AS OF. DESCRIBE HISTORY. Rollback after bad MERGE. Retention tied to VACUUM (default 7 days). Don't vacuum if you need long time travel.",
        "https://docs.databricks.com/en/delta/history.html",
    ),
    "Schema evolution & enforcement": (
        "High", "Completed", "2023-06-01", "2025-06-01",
        "SAY: Enforcement rejects extra/wrong types by default. Evolution: mergeSchema / spark.databricks.delta.schema.autoMerge. Auto Loader schemaLocation tracks drift. I allow additive columns, fail on type widening unless planned.",
        "https://docs.databricks.com/en/delta/update-schema.html",
    ),
    "Notebooks basics & widgets": (
        "High", "Completed", "2022-06-01", "2023-06-01",
        "SAY: Cells SQL/Python/Scala. Widgets parameterize env, load_date. dbutils.widgets.get. Production: jobs pass parameters; notebooks stay thin, logic in repos.",
        "https://docs.databricks.com/en/notebooks/index.html",
    ),
    "Cluster types (All-purpose, Job, SQL Warehouse)": (
        "High", "Completed", "2022-06-01", "2024-01-01",
        "SAY: All-purpose = interactive, always-on cost. Job cluster = dedicated per run, cheaper, isolated. SQL Warehouse = BI/JDBC. I develop on all-purpose, production on job clusters.",
        "https://docs.databricks.com/en/compute/index.html",
    ),
    "Cluster sizing & autoscaling": (
        "High", "Completed", "2023-01-01", "2025-06-01",
        "SAY: Size for shuffle (more workers) vs memory (bigger workers). Min/max autoscaling. Spot for non-critical. Photon for SQL/Delta. Watch idle all-purpose clusters — biggest cost leak.",
        "https://docs.databricks.com/en/compute/cluster-config-best-practices.html",
    ),
    "OPTIMIZE (compaction / Z-Ordering)": (
        "High", "Completed", "2023-06-01", "2025-06-01",
        "SAY: OPTIMIZE compact small files. ZORDER BY (filter columns) co-locates data for data skipping. Don't Z-order 10 columns — 1–4 high-filter cols. Schedule after MERGE.",
        "https://docs.databricks.com/en/delta/optimize.html",
    ),
    "VACUUM": (
        "High", "Completed", "2023-06-01", "2025-06-01",
        "SAY: Deletes unreferenced files older than retention (default 168h). Needed after OPTIMIZE/MERGE or storage grows. Time travel limited after vacuum. DRY RUN first.",
        "https://docs.databricks.com/en/delta/vacuum.html",
    ),
    "Liquid Clustering": (
        "Medium", "In Progress", "2025-06-01", None,
        "SAY: Replaces static Hive-style partitions/ZORDER for many tables. CLUSTER BY keys; incremental clustering. Better when partition columns were guessed wrong. Mention as modern alternative in interviews.",
        "https://docs.databricks.com/en/delta/clustering.html",
    ),
    "Bucketing": (
        "Low", "In Progress", "2024-01-01", None,
        "SAY: Hive buckets hash-pre-shuffle joins. Rarely my first tool on Databricks vs ZORDER/liquid/AQE. Know the term; don't oversell.",
        "https://spark.apache.org/docs/latest/sql-data-sources-load-save-functions.html#bucketing-sorting-and-partitioning",
    ),
    "Predictive Optimization": (
        "Medium", "In Progress", "2025-09-01", None,
        "SAY: UC managed tables: Databricks automatically OPTIMIZE/VACUUM. I still know manual OPTIMIZE for non-managed or when I need a specific ZORDER.",
        "https://docs.databricks.com/en/optimizations/predictive-optimization.html",
    ),
    "Catalyst Optimizer overview": (
        "High", "Completed", "2023-06-01", "2025-01-01",
        "SAY: Rule-based + cost-based optimizer: analysis → logical optimization (predicate pushdown, column pruning) → physical planning → codegen. UDFs block many optimizations.",
        "https://www.databricks.com/glossary/catalyst-optimizer",
    ),
    "Logical Plan": (
        "Medium", "Completed", "2023-06-01", "2025-01-01",
        "SAY: What to do: parsed SQL/DF unresolved → analyzed (catalog types) → optimized logical. df.explain('extended') shows Parsed/Analyzed/Optimized Logical.",
        "https://spark.apache.org/docs/latest/sql-performance-tuning.html",
    ),
    "Physical Plan": (
        "Medium", "Completed", "2023-06-01", "2025-01-01",
        "SAY: How: FileScan, BroadcastHashJoin, SortMergeJoin, HashAggregate, Shuffle. WholeStageCodegen. Compare with Spark UI SQL tab.",
        "https://spark.apache.org/docs/latest/sql-performance-tuning.html",
    ),
    "Adaptive Query Execution (AQE) in Databricks": (
        "High", "Completed", "2024-01-01", "2025-12-01",
        "SAY: On by default. Runtime coalescing, dynamic join strategy, skew handling. Demo: join that becomes broadcast after stats. Don't disable without reason.",
        "https://docs.databricks.com/en/optimizations/aqe.html",
    ),
    "Lakeflow Declarative Pipelines (DLT) concepts": (
        "Medium", "In Progress", "2025-09-01", None,
        "SAY: Declarative medallion: @dlt.table, expectations (warn/drop/fail). Streaming + batch unified. Replaces some handwritten MERGE orchestration. Also called DLT; rebranded Lakeflow Declarative Pipelines.",
        "https://docs.databricks.com/en/delta-live-tables/index.html",
    ),
    "Lakeflow Jobs orchestration": (
        "Medium", "In Progress", "2025-09-01", None,
        "SAY: Workflows/Jobs: multi-task DAG, retries, alerts, job clusters. ADF can still trigger Databricks jobs via API. For Databricks-native, Jobs > dbutils.notebook.run in prod.",
        "https://docs.databricks.com/en/workflows/jobs/index.html",
    ),
    "Bronze layer concepts": (
        "High", "Completed", "2022-06-01", "2023-12-01",
        "SAY: Raw, append-only, source-aligned. Keep original payload + audit cols (load_ts, source_file, batch_id). No business cleanses. Replayable.",
        "https://www.databricks.com/glossary/medallion-architecture",
    ),
    "Silver layer concepts": (
        "High", "Completed", "2022-06-01", "2024-06-01",
        "SAY: Cleaned, typed, deduped, conformed. SCD2 for entities. DQ rules. Joinable enterprise tables. MERGE from Bronze using business keys.",
        "https://www.databricks.com/glossary/medallion-architecture",
    ),
    "Gold layer concepts": (
        "High", "Completed", "2022-06-01", "2024-06-01",
        "SAY: Business-ready stars/aggregates, KPIs, marts for Power BI/Synapse. Grain documented. Incremental refresh friendly. Conformed dims.",
        "https://www.databricks.com/glossary/medallion-architecture",
    ),
    "Auto Loader / incremental ingestion patterns": (
        "High", "Completed", "2023-06-01", "2025-06-01",
        "SAY: cloudFiles source: incrementally ingest new files, schema inference/evolution, RocksDB/JSON checkpoint. Better than directory listing. Use with notification mode at scale.",
        "https://docs.databricks.com/en/ingestion/auto-loader/index.html",
    ),
    "Batch vs streaming ingestion in Databricks": (
        "High", "Completed", "2023-06-01", "2025-06-01",
        "SAY: Batch: scheduled read of a date partition. Streaming: Auto Loader / Event Hub continuous. Same Delta MERGE into Silver. Trigger AvailableNow = batch-like streaming for daily loads.",
        "https://docs.databricks.com/en/structured-streaming/index.html",
    ),
    "Unity Catalog (governance, access control, lineage)": (
        "High", "In Progress", "2025-01-01", None,
        "SAY: Three-level namespace catalog.schema.table. GRANT on catalogs/schemas/tables. Lineage in UI. Replaces hive_metastore for new work. Column masking for PII. Interview must-mention for 2026 Databricks roles.",
        "https://docs.databricks.com/en/data-governance/unity-catalog/index.html",
    ),
    "Databricks Workflows / Jobs orchestration": (
        "High", "Completed", "2023-06-01", "2025-06-01",
        "SAY: Multi-task job: Bronze notebook → Silver → Gold with dependencies. Retries, timeout, email/webhook. Trigger from ADF or schedule. YAML/DABs for CI.",
        "https://docs.databricks.com/en/workflows/index.html",
    ),
    "Delta Live Tables expectations (data quality rules)": (
        "Medium", "In Progress", "2025-06-01", None,
        "SAY: expect(), expect_or_drop(), expect_or_fail() on constraints (not null, valid codes). Metrics in event log. Without DLT I do the same with filter + quarantine table + log counts.",
        "https://docs.databricks.com/en/delta-live-tables/expectations.html",
    ),
    "Repos & Git integration / CI-CD basics": (
        "High", "Completed", "2023-06-01", "2025-12-01",
        "SAY: Databricks Repos ↔ Azure DevOps/GitHub. Feature branch, PR, deploy via Azure DevOps pipeline or Databricks Asset Bundles to target workspace. Never edit prod notebooks by hand.",
        "https://docs.databricks.com/en/repos/index.html",
    ),
    "Cluster Policies & Photon engine": (
        "Medium", "In Progress", "2025-01-01", None,
        "SAY: Policies restrict instance types/cost for governance. Photon = C++ vectorized engine for SQL/Delta, faster scans/joins, extra DBU. I enable Photon on job clusters when SQL-heavy.",
        "https://docs.databricks.com/en/compute/photon.html",
    ),
    "Delta Sharing": (
        "Low", "In Progress", "2025-06-01", None,
        "SAY: Open protocol to share Delta tables with external orgs without copying. Unity Catalog managed shares. Alternative to extracting CSV for partners.",
        "https://docs.databricks.com/en/data-sharing/index.html",
    ),
    "Secrets management (Key Vault integration)": (
        "High", "Completed", "2023-01-01", "2024-12-01",
        "SAY: Never hardcode secrets. Databricks secret scope backed by Azure Key Vault. dbutils.secrets.get. ADF linked services use KV too. Rotate passwords without changing code.",
        "https://docs.databricks.com/en/security/secrets/secret-scopes.html",
    ),
}

PROJ = {
    "SQL Server -> ADF -> ADLS Gen2 -> Databricks -> Bronze -> Silver -> Gold": (
        "High", "Completed", "2022-06-01", "2025-06-01",
        "YOUR STORY (Amdocs / US Cellular): Oracle Order Management + billing tables → extract (cron/Python or ADF Copy + SHIR) → land raw → clean/reconcile stuck orders → serve ops/BI. Domain: telecom orders, backlog, SLA. YOU owned: reconciliation ETL, PL/SQL bulk fixes, HTML run reports. METRICS YOU MAY SAY: ~60% less manual recon, 50%+ faster critical SQL. Do NOT invent '50 source tables' as your count — that is an interview design question.",
        "https://learn.microsoft.com/azure/data-factory/tutorial-hybrid-copy-powershell",
    ),
    "ADF pipeline design & linked services": (
        "High", "Completed", "2022-06-01", "2025-01-01",
        "MAP TO USCC: Linked service = Oracle (SHIR if on-prem) + ADLS + Databricks + Key Vault. Dataset parameterized by table name. Your real analog: Oracle connection + Python extract config, not 50 ADF pipelines. Azure interview: one template pipeline, many linked datasets.",
        "https://learn.microsoft.com/azure/data-factory/concepts-linked-services",
    ),
    "ADF orchestration (triggers, pipelines, activities)": (
        "High", "Completed", "2022-06-01", "2025-01-01",
        "YOUR ANALOG: Linux cron + Python jobs + Sendmail reports (Amdocs). Azure words: schedule trigger = cron; Copy = extract; Stored Proc = PL/SQL bulk fix; Notebook = transform; If/ForEach = config-driven table list. Triggers: daily batch for orders, not streaming unless asked.",
        "https://learn.microsoft.com/azure/data-factory/concepts-pipelines-activities",
    ),
    "Landing raw data into ADLS Gen2": (
        "High", "Completed", "2022-06-01", "2025-01-01",
        "SAY for Azure: /bronze/uscc/orders/yyyy/mm/dd/ immutable. YOUR analog: raw extracts / staging tables before recon. Keep source payload. Partition by business date for prune. Parquet over CSV once past landing.",
        "https://learn.microsoft.com/azure/storage/blobs/data-lake-storage-introduction",
    ),
    "Full Load implementation": (
        "High", "Completed", "2020-06-01", "2023-01-01",
        "USCC: small ref/code tables, or first backfill of an entity. Staging then swap so ops never see half-load. Control flag load_type=full. Do not full-refresh high-volume order history daily — that is what incremental is for.",
        "https://learn.microsoft.com/azure/data-factory/copy-activity-overview",
    ),
    "Incremental Load implementation (watermarking)": (
        "High", "Completed", "2022-06-01", "2025-06-01",
        "YOUR PATTERN: last successful modified_ts / order status change. Lookup watermark → WHERE last_upd > watermark → MERGE/recon → update watermark LAST. Late orders: lookback overlap. Deletes: status flags or CDC — timestamp-only misses hard deletes. Same pattern you used for backlog delta, not a fake table count.",
        "https://learn.microsoft.com/azure/data-factory/tutorial-incremental-copy-portal",
    ),
    "Metadata-driven ingestion framework": (
        "High", "Completed", "2023-01-01", "2025-06-01",
        "INTERVIEW 50/40/10: one pipeline + control table (table_name, load_type, watermark_col, src, tgt, is_active). ForEach. Onboard = INSERT row. YOUR TRUTH: you did not publish a '50 table' number — say 'I would scale the same config pattern we used for order entities; I will not guess a table count I cannot defend.'",
        "https://learn.microsoft.com/azure/data-factory/copy-data-tool-metadata-driven",
    ),
    "Raw data ingestion into Delta (Bronze)": (
        "High", "Completed", "2022-06-01", "2025-06-01",
        "Azure: append-only Delta, source-aligned order/billing rows + ingest meta. YOUR analog: land Oracle extract before recon so you can replay a bad run. Idempotent on batch_id or source PK. Do not drop columns in Bronze.",
        "https://docs.databricks.com/en/ingestion/auto-loader/index.html",
    ),
    "Audit columns (load timestamp, source, batch id)": (
        "High", "Completed", "2022-06-01", "2023-12-01",
        "YOU already did this: logging/audit for production traceability on USCC recon. Names: src_system=USCC/Amdocs, ingestion_ts UTC, batch_id=run id, job_name. Same columns Bronze→Gold so SLA reports join.",
        "https://www.databricks.com/glossary/medallion-architecture",
    ),
    "Logging & error handling during ingestion": (
        "High", "Completed", "2022-06-01", "2025-06-01",
        "STAR you can defend: stuck/failed orders, P1/P2 with DBA/Dev. Retry transient; quarantine poison rows; HTML email of exceptions (Sendmail). Watermark not advanced on fail. Azure: Copy faultTolerance + fail activity → alert. Dynatrace/BMC Remedy = your real monitor words.",
        "https://learn.microsoft.com/azure/data-factory/copy-activity-fault-tolerance",
    ),
    "Data quality checks & validation rules": (
        "High", "Completed", "2022-06-01", "2025-06-01",
        "USCC recon IS DQ: order PK present, status in allowed set, upstream vs downstream match, backlog age vs SLA. Fail vs quarantine vs warn. Log counts. Do not claim Great Expectations unless you used it — say Python/SQL assertions + exception queue.",
        "https://docs.databricks.com/en/delta-live-tables/expectations.html",
    ),
    "Deduplication logic": (
        "High", "Completed", "2022-06-01", "2025-06-01",
        "Order entity: ROW_NUMBER() PARTITION BY order_id ORDER BY last_upd DESC, ingest_ts DESC → rn=1. Late duplicate status updates are normal in OMS. MERGE when hash of status/attrs changes. Same as your backlog uniqueness.",
        "https://docs.databricks.com/en/delta/merge.html",
    ),
    "SCD Type 2 implementation": (
        "High", "Completed", "2023-01-01", "2025-06-01",
        "If they ask SCD2 on USCC: customer/account or product/offer attrs that must keep history (plan changes). MERGE expire is_current=1 then insert new SK. Facts (orders) keep SK. If you only overwrote status in prod, say SCD1 for order status, SCD2 for customer — do not invent Type-6 hybrids.",
        "https://docs.databricks.com/en/delta/merge.html",
    ),
    "Delta MERGE for upserts": (
        "High", "Completed", "2023-01-01", "2025-06-01",
        "YOUR SQL analog: PL/SQL bulk status correction / MERGE on order_id. Azure: MERGE INTO silver USING updates ON bk. Unique source keys or MERGE fails. Update watermark after MERGE succeeds. 50%+ query win = indexing/plans on Oracle, not Delta OPTIMIZE unless you did it.",
        "https://docs.databricks.com/en/delta/merge.html",
    ),
    "Curated / aggregated business tables": (
        "High", "Completed", "2022-06-01", "2025-06-01",
        "Gold for USCC ops: daily backlog by status, SLA-breach counts, exception codes — the HTML report grain. Not 'store sales'. Document KPI with business (what is stuck). BI/self-service is a partner, not your fake Tableau dashboard unless you built it.",
        "https://learn.microsoft.com/power-bi/guidance/star-schema",
    ),
    "Star schema modeling in Gold layer": (
        "High", "Completed", "2023-01-01", "2025-06-01",
        "Draw: FactOrderEvent (grain=one order line/status change) + DimDate + DimCustomer + DimProduct/Plan + DimStatus. Integer SKs. Degenerate: order_number on the fact. This is how you answer modelling even if day-job was recon tables.",
        "https://learn.microsoft.com/power-bi/guidance/star-schema",
    ),
    "End-to-end error handling strategy": (
        "High", "Completed", "2022-06-01", "2025-12-01",
        "STAR: production defect / 70% copy fail analog = job dies mid-batch. You: idempotent rerun from last watermark, staging, exception queue, P1 bridge. Result: faster MTTR, Client Appreciation / EOM. Azure: retry + don't update watermark + date-slice reload.",
        "https://learn.microsoft.com/azure/data-factory/copy-activity-fault-tolerance",
    ),
    "Centralized logging framework": (
        "Medium", "Completed", "2022-06-01", "2025-06-01",
        "YOU: Python logging + audit tables + HTML email + Dynatrace + Remedy tickets. Azure names: pipeline_run_log (run_id, entity, rows, status, error) → Log Analytics. One SLA view: last success, lag, fail activity.",
        "https://learn.microsoft.com/azure/data-factory/monitor-using-azure-monitor",
    ),
    "Audit column standards across layers": (
        "High", "Completed", "2022-06-01", "2024-06-01",
        "Team standard you can claim: src_system, batch_id, ingestion_ts on every landing table (you built this for compliance/traceability). Silver adds quality_flag. Gold keeps batch_id for lineage, drops raw payload.",
        "https://www.databricks.com/glossary/medallion-architecture",
    ),
    "Schema drift / schema evolution handling": (
        "High", "Completed", "2023-06-01", "2025-12-01",
        "OMS reality: new status codes / extra XML-JSON fields from Flask APIs. Land extra cols in Bronze (rescue/unknown). Silver mapping via PR. Type changes are breaking. Do not claim Auto Loader schemaLocation unless you used it.",
        "https://learn.microsoft.com/azure/data-factory/copy-activity-schema-and-type-mapping",
    ),
    "Parameterization & dynamic pipelines in ADF": (
        "High", "Completed", "2022-06-01", "2025-06-01",
        "YOUR analog: argparse/config for entity, load_date, env. Azure: pipeline params p_table, p_load_date; dataset folder @concat('bronze/uscc/', pipeline().parameters.p_table). One pipeline, many order-related entities.",
        "https://learn.microsoft.com/azure/data-factory/how-to-use-parameters",
    ),
    "Monitoring & alerting (Azure Monitor, Log Analytics)": (
        "High", "Completed", "2022-06-01", "2025-12-01",
        "REAL words first: Dynatrace APM, BMC Remedy, regex log alerts, paramiko SSH, email. Azure translation: Monitor + Log Analytics on ADF/Databricks fails, alert on duration vs SLA and watermark lag. Tata Communications: 99.9%+ availability story is NOC, not ADF — only use if they ask earlier role.",
        "https://learn.microsoft.com/azure/data-factory/monitor-using-azure-monitor",
    ),
    "Secrets management (Azure Key Vault) in ADF/Databricks": (
        "High", "Completed", "2023-01-01", "2024-12-01",
        "SAY: no passwords in Git/cron scripts (you should have moved off hardcoded). Azure: KV linked service, Databricks secret scope, managed identity. Oracle creds in KV. If current job still uses env files, say you want KV and do not pretend it is live.",
        "https://learn.microsoft.com/azure/data-factory/store-credentials-in-key-vault",
    ),
    "CI/CD for data pipelines (Azure DevOps)": (
        "High", "In Progress", "2024-01-01", None,
        "HONEST: Git/GitHub + code review is on your resume; Azure DevOps ADF publish is the interview target, not a number you invented. Say: feature branch, PR, deploy Dev→Test→Prod, env params in KV. Do not claim DABs unless you used them.",
        "https://learn.microsoft.com/azure/data-factory/continuous-integration-delivery",
    ),
    "Unit/data testing for pipelines": (
        "Medium", "In Progress", "2025-01-01", None,
        "HONEST GAP: pytest/unittest on resume; prod DQ was recon counts + exception queue. Interview: one pytest on a recon function + row-count check vs source. Do not claim chispa/Great Expectations.",
        "https://learn.microsoft.com/azure/data-factory/continuous-integration-delivery-improvements",
    ),
}

INT = {
    "Branching, merge/rebase, pull requests": (
        "High", "Completed", "2021-06-01", "2023-06-01",
        "SAY: feature/xxx from main, PR, squash merge. Rebase to update branch; avoid rebase of shared branches. ADF/Databricks artifacts in Git. Conflict resolution on JSON pipelines needs care.",
        "https://learn.microsoft.com/devops/develop/git/git-branching-guidance",
    ),
    "Python fundamentals (functions, OOP basics, list/dict comprehensions)": (
        "High", "Completed", "2021-03-01", "2023-01-01",
        "SAY: Functions, typing, pathlib, dataclasses. Comprehensions for small in-memory work — not for Spark data. Interview live-code: parse JSON, group dicts, write a watermark helper.",
        "https://docs.python.org/3/tutorial/",
    ),
    "Working with APIs, JSON, file I/O": (
        "High", "Completed", "2022-01-01", "2024-01-01",
        "SAY: Flask/FastAPI JSON/XML sync between USCC upstream/downstream OMS (your resume). Pagination + retries. Land API JSON in Bronze then explode. Rate limits. Do not invent Kafka unless asked as design.",
        "https://docs.python.org/3/library/json.html",
    ),
    "Azure Storage (Blob/ADLS Gen2) basics": (
        "High", "Completed", "2021-09-01", "2023-01-01",
        "SAY: Blob vs ADLS Gen2 (HNS, POSIX ACLs, directories). abfss://container@account.dfs.core.windows.net/path. Auth: account key (dev), SAS, managed identity (prod).",
        "https://learn.microsoft.com/azure/storage/blobs/data-lake-storage-introduction",
    ),
    "Azure SQL / Synapse basics": (
        "High", "Completed", "2022-01-01", "2024-06-01",
        "SAY: Azure SQL = OLTP/serving. Synapse dedicated = MPP warehouse (distributions, DWUs). Serverless SQL over ADLS. Fabric Warehouse similar idea. I use SQL for Gold serving or stored-proc ETL.",
        "https://learn.microsoft.com/azure/synapse-analytics/sql-data-warehouse/sql-data-warehouse-overview-what-is",
    ),
    "Azure Key Vault & IAM/RBAC basics": (
        "High", "Completed", "2023-01-01", "2025-01-01",
        "SAY: RBAC roles (Blob Data Contributor, SQL DB contributor) vs access keys. MI on ADF/Databricks. Least privilege. KV for secrets, RBAC for data plane.",
        "https://learn.microsoft.com/azure/key-vault/general/overview",
    ),
    "Data governance concepts (lineage, cataloging)": (
        "High", "In Progress", "2024-06-01", None,
        "SAY: Catalog = what tables exist + owners. Lineage = Bronze file → Silver → Gold → report. Unity Catalog / Purview. Impact analysis before dropping a column.",
        "https://learn.microsoft.com/azure/purview/overview",
    ),
    "Data masking / PII handling": (
        "High", "Completed", "2023-06-01", "2025-06-01",
        "SAY: Classify PII in Bronze, tokenize/hash/mask in Silver. Restrict Gold. UC column masks / dynamic data masking in SQL. No PII in logs. Access via groups not personal accounts.",
        "https://learn.microsoft.com/azure/azure-sql/database/dynamic-data-masking-overview",
    ),
    "Encryption at rest & in transit": (
        "Medium", "Completed", "2023-01-01", "2024-12-01",
        "SAY: At rest: Azure storage SSE, SQL TDE. In transit: TLS/HTTPS, encryption in transit on Spark. CMK in KV for extra control. Don't claim you 'implemented AES' unless you did.",
        "https://learn.microsoft.com/azure/security/fundamentals/encryption-overview",
    ),
    "Unit testing for data pipelines (pytest, chispa, etc.)": (
        "Medium", "In Progress", "2025-01-01", None,
        "SAY: Pure Python functions tested with pytest. Spark transforms: chispa assert_df_equality on tiny DFs. CI runs on sample. I still rely more on DQ counts in prod — closing that gap.",
        "https://docs.pytest.org/",
    ),
    "CI/CD pipelines for data engineering (Azure DevOps/GitHub Actions)": (
        "High", "Completed", "2024-01-01", "2025-12-01",
        "SAY: Build validates; release deploys ADF ARM + Databricks bundle to env. Secrets from KV. Approvals for prod. Idempotent deploys.",
        "https://learn.microsoft.com/azure/data-factory/continuous-integration-delivery",
    ),
    "Pipeline monitoring & alerting": (
        "High", "Completed", "2023-06-01", "2025-12-01",
        "SAY: Success/failure, duration vs SLA, rows loaded vs 7-day avg, watermark lag. Alerts to Teams/email. First debug: activity output error + input parameters + last good run.",
        "https://learn.microsoft.com/azure/data-factory/monitor-visually",
    ),
    "Cost optimization (cluster sizing, spot instances, autoscaling)": (
        "High", "Completed", "2024-01-01", "2025-12-01",
        "SAY: Job clusters not 24/7 all-purpose. Auto-terminate. Spot + fallback. Photon when it pays back. Compact small files. Incremental not full. Lifecycle policy on ADLS logs. Partition pruning.",
        "https://learn.microsoft.com/azure/databricks/administration-guide/capacity/plan-optimize-cost",
    ),
    "System design for data pipelines (whiteboard a pipeline end-to-end)": (
        "High", "In Progress", "2025-06-01", None,
        "DRAW USCC: Oracle OMS/billing → ingest (ADF Copy/SHIR or Python) → Bronze orders → Silver recon (PK, status, SLA) → Gold backlog star → ops report/PBI. Add watermark, retries, DQ, Dynatrace/Monitor, IAM, cost. 50 tables = control table design, not your headcount of tables.",
        "https://learn.microsoft.com/azure/architecture/example-scenario/data/hybrid-etl-with-adf",
    ),
    "Behavioral questions (STAR method) for DE roles": (
        "High", "In Progress", "2025-06-01", None,
        "STAR from YOUR resume: stuck-order backlog (60% less manual), slow SQL (50%+ via index/plan), P1 with DBA (Client Appreciation), Tata NOC 99.9%+. 70% copy fail = idempotent rerun story. Never say 50 tables were yours.",
        "https://learn.microsoft.com/azure/data-factory/tutorial-incremental-copy-overview",
    ),
}

SHEETS = {
    "SQL": SQL,
    "Data Warehouse": DWH,
    "PySpark": PYSPARK,
    "Databricks": DBX,
    "Project": PROJ,
    "Interview Essentials": INT,
}

PRACTICE = [
    ("2026-08-18", 8, "Joins, anti-join NOT EXISTS", "Medium", "LeetCode SQL", 4, "50-table ADF DESIGN question warmup — not your table count"),
    ("2026-08-19", 7, "ROW_NUMBER / RANK / DENSE_RANK", "Medium", "HackerRank", 4, "Dedup pattern rn=1"),
    ("2026-08-20", 6, "LEAD/LAG, running totals", "Medium", "StrataScratch", 3, "Need more frame-clause practice"),
    ("2026-08-21", 9, "CTE, MERGE, SCD2 sketch", "Hard", "SSMS + notes", 4, "Wrote MERGE expire+insert"),
    ("2026-08-22", 5, "Indexes, sargable predicates", "Medium", "Notes 11 interview round", 4, "PK vs unique, seek vs scan"),
    ("2026-08-23", 8, "Window + gaps and islands", "Hard", "LeetCode", 3, "Revise islands pattern"),
    ("2026-08-24", 6, "Incremental watermark SQL", "Medium", "Azure Learn + own project", 4, "Overlap lookback for late rows"),
]

HOW_TO = [
    ("Intro (60–90s)", "Tell me about yourself",
     "Bhushan Jain, ~5 years, Pune. Amdocs (Jun 2022–present) on US Cellular telecom order management: Python + Oracle SQL/PL-SQL + Linux cron ETL, Flask/FastAPI integrations, production support. Earlier Tata Communications (Jun 2020–May 2022) NOC/Unix, 99.9%+ availability, 10+ Python/shell automations (~35% productivity). Strongest: SQL + Python. Azure DE wording: I ingest/reconcile order+billing data and would land it ADF→ADLS→Databricks medallion. Lead metric: ~60% less manual recon and 50%+ faster critical SQL. Stop. Do not list Snowflake/life-sciences/OBS from the ATS resume unless you can defend them.",
     "NOW → employer/client → stack you can defend → ONE metric → stop."),
    ("Project story", "Walk through your architecture",
     "USCC OMS/billing on Oracle. Extract changed orders (watermark/status). Stage raw. Recon: PK, allowed status, upstream vs downstream. Bulk PL/SQL fixes. HTML exception report via cron/Sendmail. Monitor Dynatrace + Remedy. Azure mapping: Oracle--SHIR/ADF Copy-->ADLS Bronze-->Databricks Silver MERGE-->Gold FactOrderEvent star-->PBI/ops. You owned recon ETL and SQL tuning, not the whole cloud account.",
     "Draw boxes. Label what YOU built vs platform team."),
    ("50 tables ADF", "40 incremental + 10 full daily",
     "This is a DESIGN question, not your resume count. Answer: one parameterized pipeline + control table (load_type, watermark_col, src, tgt, is_active). ForEach. Full = small dims; incremental = order/billing facts. Onboard = INSERT row. If they ask how many tables YOU had: 'I will not quote a number I cannot defend; the pattern is what I would use.'",
     "Never say you personally onboarded 50 USCC tables."),
    ("Copy 70% failed", "How do you recover?",
     "Map to a mid-batch USCC job fail. Idempotent MERGE/staging; do not advance watermark; retry transient; quarantine poison; HTML exception list; P1 with DBA. Re-run from last watermark or delete that date slice then reload. Result: faster MTTR, Client Appreciation story.",
     "Use YOUR incident, not a fake ADF screenshot."),
    ("PK vs Unique", "Difference?",
     "PK: one per table, no NULLs, identifies row, often clustered. UNIQUE: multiple allowed, one NULL in SQL Server unique index (historically), alternate keys. Both enforce uniqueness.",
     "Give an example: email UNIQUE, customer_id PK."),
    ("RANK trio", "ROW_NUMBER vs RANK vs DENSE_RANK",
     "Same OVER clause. RN unique. RANK gaps after ties. DENSE no gaps. Dedup = RN=1. Top salary per dept = RN or DENSE depending if ties should share rank.",
     "Write OVER(PARTITION BY dept ORDER BY sal DESC) on a whiteboard."),
    ("LEFT vs INNER", "Can LEFT be replaced by INNER?",
     "Only if you don't need unmatched left rows. WHERE on right table makes LEFT behave as INNER. Put right filters in ON to keep true LEFT.",
     "Draw two circles."),
    ("CTE vs table vs view", "When each?",
     "CTE: readable query-scoped. Temp table: materialize + index. View: reusable saved SELECT, no data. Cannot always UPDATE CTE; yes if single base table.",
     "Mention recursive CTE separately."),
    ("Incremental pipeline", "How do you load incrementally?",
     "Watermark column (modified_ts) in control table. Read last value → filter source → land Bronze → MERGE Silver → write new watermark. CDC if deletes required. Lookback window for late data.",
     "Mention failure = don't update watermark."),
    ("Spark join tuning", "How do you optimize a slow join?",
     "Check sizes: broadcast small dim. Skew: AQE / salt / isolate hot key. Partition/ZORDER join keys. Avoid UDFs in join keys. Spark UI stragglers. Repartition by join key if both large.",
     "Always mention Spark UI evidence."),
    ("Delta vs Parquet", "Why Delta?",
     "Parquet = files. Delta = Parquet + transaction log: ACID, MERGE, time travel, schema enforcement. Lakehouse table format.",
     "One sentence then table."),
    ("Medallion", "Bronze Silver Gold?",
     "Bronze raw replayable. Silver cleaned conformed. Gold products/stars. Quality increases, volume often decreases (aggregates).",
     "Map to their company nouns."),
    ("Narrow vs wide", "Spark transformations",
     "Narrow: no shuffle (filter, map). Wide: shuffle (groupBy, join). Cost lives in wide ops.",
     "Tie to DAG stages."),
    ("SCD2", "How implemented?",
     "Surrogate key + business key + hash + is_current + dates. MERGE expire+insert. Facts use SK.",
     "Don't forget deletes/is_active."),
    ("System design 5 min", "Design a telecom order analytics pipeline",
     "Start from YOUR domain: Oracle OMS CDC/watermark → ADF/Python ingest → ADLS Bronze → Spark Silver recon (order PK, status, SLA) → Gold FactOrderEvent + DimCustomer/Plan/Status → ops report/PBI. Cover late orders, PII (MSISDN), grants, job cluster vs cron, Dynatrace/Monitor, cost. Ask volume/SLA before drawing.",
     "Same boxes as USCC, not a generic retail demo."),
    ("Behavioral", "Tell me about a failure or win",
     "Pick one STAR only: (1) stuck-order recon → Python/PL-SQL + reports → ~60% less manual. (2) slow critical SQL → index/plan → 50%+. (3) P1 with DBA → Client Appreciation. (4) Tata NOC automations → ~35% / 99.9%+. Learning: idempotent rerun, don't move watermark on fail.",
     "90 seconds out loud. I not we."),
]


def parse_d(s):
    if not s:
        return None
    y, m, d = map(int, s.split("-"))
    return date(y, m, d)


def fill_topic_sheet(ws, mapping):
    for row in range(5, ws.max_row + 1):
        sub = ws.cell(row, 3).value
        if not sub or sub not in mapping:
            continue
        pr, st, ds, dc, notes, link = mapping[sub]
        ws.cell(row, 4).value = pr
        ws.cell(row, 5).value = st
        ws.cell(row, 6).value = parse_d(ds)
        ws.cell(row, 6).number_format = "YYYY-MM-DD"
        if dc:
            ws.cell(row, 7).value = parse_d(dc)
            ws.cell(row, 7).number_format = "YYYY-MM-DD"
        else:
            ws.cell(row, 7).value = None
        ws.cell(row, 8).value = notes
        ws.cell(row, 8).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 9).value = link
        # light fills by status
        fills = {
            "Completed": PatternFill("solid", fgColor="C6EFCE"),
            "In Progress": PatternFill("solid", fgColor="FFEB9C"),
            "Revise Again": PatternFill("solid", fgColor="FFC7CE"),
        }
        if st in fills:
            ws.cell(row, 5).fill = fills[st]
        pr_fill = {"High": PatternFill("solid", fgColor="F4B183"), "Medium": PatternFill("solid", fgColor="BDD7EE"), "Low": PatternFill("solid", fgColor="D9D9D9")}
        if pr in pr_fill:
            ws.cell(row, 4).fill = pr_fill[pr]
        ws.row_dimensions[row].height = 48
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["H"].width = 72
    ws.column_dimensions["I"].width = 45


def add_how_to_sheet(wb):
    if "How_to_Answer" in wb.sheetnames:
        del wb["How_to_Answer"]
    ws = wb.create_sheet("How_to_Answer", 1)
    title = Font(bold=True, size=16, color="FFFFFF")
    header = Font(bold=True, color="FFFFFF")
    ws.merge_cells("A1:D1")
    ws["A1"] = "How to answer every tracker topic — 5 years Data Engineer (Azure / ADF / Databricks / SQL)"
    ws["A1"].font = title
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws.merge_cells("A2:D2")
    ws["A2"] = "Bhushan Jain | Amdocs–US Cellular (Jun 2022–present) | Tata Communications (Jun 2020–May 2022). Speak Oracle+Python+cron first; map to ADF/ADLS/Databricks when they ask Azure. Never quote 50 tables as your inventory."
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 36
    headers = ["Theme", "They ask", "What you say (5 YOE)", "Delivery tip"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i, h)
        c.font = header
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(wrap_text=True)
    for r, row in enumerate(HOW_TO, 5):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 70
    # universal formula
    ws.merge_cells("A22:D22")
    ws["A22"] = "UNIVERSAL ANSWER FORMULA (use on any topic in SQL / DWH / PySpark / Databricks / Project tabs)"
    ws["A22"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A22"].fill = PatternFill("solid", fgColor="1F4E79")
    ws.merge_cells("A23:D26")
    ws["A23"] = (
        "1) NAME IT: 'X is …' (precise definition).\n"
        "2) PLACE IT: 'In my pipeline it sits in Bronze/Silver/Gold / ADF activity / Spark stage …'.\n"
        "3) DO IT: one concrete mechanism (watermark column, MERGE ON bk, broadcast hint, ZORDER col).\n"
        "4) COST/RISK: shuffle, small files, NOT IN + NULL, LEFT+WHERE becomes INNER, vacuum vs time travel.\n"
        "5) PROOF: Spark UI / row counts / SLA minutes / onboarding a table via config.\n"
        "6) If you are weak on a row marked In Progress: say 'I have used the adjacent pattern in prod; this exact feature I am hands-on this week' — then still give a correct definition. Never bluff Photon/DLT internals."
    )
    ws["A23"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[23].height = 20
    ws.row_dimensions[24].height = 20
    ws.row_dimensions[25].height = 20
    ws.row_dimensions[26].height = 40
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 88
    ws.column_dimensions["D"].width = 42
    ws.freeze_panes = "A5"


def add_my_story_sheet(wb):
    if "My_Story" in wb.sheetnames:
        del wb["My_Story"]
    ws = wb.create_sheet("My_Story", 1)
    title = Font(bold=True, size=16, color="FFFFFF")
    header = Font(bold=True, color="FFFFFF")
    ws.merge_cells("A1:C1")
    ws["A1"] = "Bhushan Jain — spoken project story (only numbers from your resume)"
    ws["A1"].font = title
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws.merge_cells("A2:C2")
    ws["A2"] = "Use this tab in interviews. Column B is what you say. Column C is what you must not invent."
    ws["A2"].alignment = Alignment(wrap_text=True)
    headers = ["Block", "Say this", "Do not say"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(4, i, h)
        c.font = header
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(wrap_text=True)
    rows = [
        ("Who you are",
         "Bhushan Jain, Pune. About five years in telecom data/ops. Current: Amdocs, client US Cellular, Jun 2022–present. Previous: Tata Communications NOC, Jun 2020–May 2022. SQL and Python are strongest.",
         "Do not say 5.9 years Azure DE at TCL overlapping Amdocs. Do not quote 9 LPA unless you want that on the table."),
        ("Current project",
         "US Cellular order management. Oracle holds orders/billing. I built Python + PL/SQL jobs on Linux cron: find stuck/failed orders, reconcile upstream vs downstream, bulk-fix status, HTML email of exceptions. Flask/FastAPI for JSON/XML sync. Dynatrace + BMC Remedy for prod.",
         "Do not claim you own the whole Azure subscription, Snowflake, or a life-sciences lake."),
        ("Azure mapping (when they ask DE)",
         "Same pipeline in Azure words: Oracle --SHIR/ADF Copy--> ADLS Bronze --Databricks MERGE--> Silver recon --Gold star FactOrderEvent--> Power BI/ops. I speak ADF as the orchestrator equivalent of cron+config, not as a tool I used all day unless the JD is Azure-only and you have practiced the UI.",
         "Do not say you ran Photon, DLT, Unity Catalog, or 50 ADF pipelines in prod."),
        ("Metrics you may use",
         "~60% less manual order recon. 50%+ faster critical Oracle SQL (index + plan). Tata: 99.9%+ availability, 10+ Python/shell automations, ~35% productivity. Awards: Employee of the Month, USCC Client Appreciation.",
         "Do not invent TB/day, 50 tables, 3-hour-to-50-minute lake SLA, or row counts you never measured."),
        ("50 tables question",
         "Design answer: one parameterized pipeline + control table (load_type full vs incremental, watermark column, source, target). 10 small dims full daily, 40 facts incremental. Onboard = INSERT config. If they ask MY count: I will not guess; we grouped order-related entities the same way.",
         "Never: I ingested 50 USCC tables."),
        ("Copy failed at 70%",
         "Job died mid-batch. Staging or MERGE by order_id so rerun is safe. Do not advance watermark. Retry transient. Quarantine poison rows. Send exception HTML. P1 with DBA. Reload from last watermark or that date slice.",
         "Do not describe an ADF monitor screen you have not used."),
        ("STAR 1 — recon",
         "S: stuck orders missed SLA. T: cut manual queue. A: I wrote Python/pandas + PL/SQL bulk status + cron report. R: ~60% less manual recon, EOM/client award.",
         "Do not credit the whole OMS rewrite to yourself."),
        ("STAR 2 — SQL",
         "S: critical recon query too slow for the batch window. T: cut runtime. A: I read the plan, added supporting index, rewrote non-sargable predicates. R: 50%+ faster.",
         "Do not say Spark AQE if it was Oracle."),
        ("STAR 3 — Tata",
         "S: repetitive NOC checks. T: reduce toil. A: I automated 10+ log/disk/threshold jobs in Python/shell. R: ~35% productivity, 99.9%+ availability with the team.",
         "Do not rebrand Tata as TCL Azure Databricks."),
        ("What you owned vs team",
         "I owned recon jobs, SQL tuning, exception reports, API glue, incident RCA. Platform/DBA owned Oracle HA. BI owned dashboards if they exist. Be explicit.",
         "Do not say I architected the medallion lakehouse for USCC."),
    ]
    for r, row in enumerate(rows, 5):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 72
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 88
    ws.column_dimensions["C"].width = 52
    ws.freeze_panes = "A5"


def fill_practice(ws):
    for i, rec in enumerate(PRACTICE):
        r = 5 + i
        ws.cell(r, 1).value = parse_d(rec[0])
        ws.cell(r, 1).number_format = "YYYY-MM-DD"
        ws.cell(r, 2).value = rec[1]
        ws.cell(r, 4).value = rec[2]
        ws.cell(r, 5).value = rec[3]
        ws.cell(r, 6).value = rec[4]
        ws.cell(r, 7).value = rec[5]
        ws.cell(r, 8).value = rec[6]


def annotate_dashboard(ws):
    ws.merge_cells("A27:F27")
    ws["A27"] = (
        "Bhushan Jain — Amdocs/US Cellular order+billing + Tata Communications NOC. "
        "Project tab + How_to_Answer + My_Story use YOUR resume metrics (~60% less manual recon, 50%+ SQL, 99.9%+ NOC, ~35% automation). "
        "50 tables is an interview DESIGN pattern, not a CV number. "
        "Yellow = Azure-only / newer Databricks — define correctly, do not bluff prod."
    )
    ws["A27"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[27].height = 48


def main():
    wb = load_workbook(SRC)
    missing = []
    for sheet, mapping in SHEETS.items():
        ws = wb[sheet]
        # verify all keys exist
        found = set()
        for row in range(5, ws.max_row + 1):
            sub = ws.cell(row, 3).value
            if sub:
                found.add(sub)
        for k in mapping:
            if k not in found:
                missing.append((sheet, k))
        fill_topic_sheet(ws, mapping)
    fill_practice(wb["Daily SQL Practice Log"])
    add_how_to_sheet(wb)
    add_my_story_sheet(wb)
    annotate_dashboard(wb["Dashboard"])
    # freeze panes on topic sheets
    for name in SHEETS:
        wb[name].freeze_panes = "A5"
        wb[name].auto_filter.ref = None
    if missing:
        raise SystemExit("Missing keys: " + str(missing))
    wb.save(OUT1)
    wb.save(OUT2)
    # also overwrite original as user asked to fill it — keep a backup first already as OUT1
    wb.save(SRC)
    print("saved", OUT1)
    print("saved", OUT2)
    print("updated original", SRC)
    # counts
    for sheet, mapping in SHEETS.items():
        from collections import Counter
        c = Counter(v[1] for v in mapping.values())
        print(sheet, dict(c), "topics", len(mapping))


if __name__ == "__main__":
    main()
